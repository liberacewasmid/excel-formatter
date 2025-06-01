import openpyxl

wb = openpyxl.load_workbook("samplebook.xlsx")
products_sheet = wb["SHEET 1"]
summary_sheet = wb["SUMMARY"]

sections = [
    {
        "start_row": 6,
        "end_row": 26,                         # Section 1
        "subtotal_row": 27,
        "summary_cell": "E3"
    },
    {
        "start_row": 29,
        "end_row": 45,                         # Section 2
        "subtotal_row": 46,
        "summary_cell": "E5"
    },
    {
        "start_row": 48,
        "end_row": 64,                         # Section 3
        "subtotal_row": 65,
        "summary_cell": "E7"
    },
    {
        "start_row": 67,
        "end_row": 83,                         # Section 4
        "subtotal_row": 84,
        "summary_cell": "E9"
    },
    {
        "start_row": 86,
        "end_row": 102,                         # Section 5
        "subtotal_row": 103,
        "summary_cell": "E11"
    },

]

qty_col = 3
price_col = 5
total_col = 6

for i, section in enumerate(sections, start=1):
    subtotal = 0
    print(f"\nFormatting Section {i} (Rows {section['start_row']} to {section['end_row']})")

    for row in range(section["start_row"], section["end_row"] + 1):
        price = products_sheet.cell(row=row, column=price_col).value
        qty = products_sheet.cell(row=row, column=qty_col).value

        print(f"Row {row}: Price={price}, Qty={qty}")

        try:
            p = float(price) if price not in [None, ""] else 0
            q = float(qty) if qty not in [None, ""] else 0

            total = p * q
            products_sheet.cell(row=row, column=total_col).value = total
            subtotal += total

        except Exception as e:
            print(f"skipping row {row} because: {e}")

    subtotal_cell = products_sheet.cell(row=section["subtotal_row"], column=total_col)
    subtotal_cell.value = subtotal

    summary_sheet[section["summary_cell"]] = subtotal
    print(f"Section {i} subtotal = {subtotal} → {section['summary_cell']}")

wb.save("samplebook_formatted.xlsx")
print("\nFormatting done. Book was saved as 'samplebook_formatted.xlsx'")
